# coding:UTF-8
# author    :ZHX、Just_silent
# init time :2021/5/12 16:08
# file      :similar.py
# IDE       :PyCharm

import heapq
import os
import random
import openpyxl

os.environ['CUDA_VISIBLE_DEVICES'] = '-1'
from bert4keras.snippets import DataGenerator, to_array
from keras.layers import Dropout, Dense
from bert4keras.optimizers import Adam
from train.utils import *
import tensorflow.compat.v1 as tf
tf.disable_v2_behavior()

from train.path import *


# 回调类
class Evaluator(keras.callbacks.Callback):
    def __init__(self, model, valid_generator, test_generator):
        self.best_val_acc = 0.
        self.model = model
        self.valid_generator = valid_generator
        self.test_generator = test_generator
    # 每迭代一次，调用一次
    def on_epoch_end(self, epoch, logs=None):
        val_acc = self.evaluate(self.valid_generator)
        if val_acc > self.best_val_acc:
            self.best_val_acc = val_acc
            self.model.save_weights(bert_base_model_save_path)
        test_acc = self.evaluate(self.test_generator)
        print(
            u'val_acc: %.5f, best_val_acc: %.5f, test_acc: %.5f\n' %
            (val_acc, self.best_val_acc, test_acc)
        )

    def evaluate(self, data):
        total, right = 0., 0.
        for x_true, y_true in data:
            y_pred = self.model.predict(x_true).argmax(axis=1)
            y_true = y_true[:, 0]
            total += len(y_true)
            right += (y_true == y_pred).sum()
        return right / total


tokenizer = Tokenizer(vocab_path, do_lower_case=True)

# 数据生成器，迭代一次返回一个batch的数据，[batch_token_ids, batch_segment_ids], batch_labels
class DataGenerator(DataGenerator):
    """数据生成器
    """
    def __iter__(self, random=False):
        batch_token_ids, batch_segment_ids, batch_labels = [], [], []
        for is_end, (text1, text2, label) in self.sample(random):
            token_ids, segment_ids = tokenizer.encode(
                text1, text2
            )
            batch_token_ids.append(token_ids)
            batch_segment_ids.append(segment_ids)
            batch_labels.append([label])
            if len(batch_token_ids) == self.batch_size or is_end:
                batch_token_ids = sequence_padding(batch_token_ids)
                batch_segment_ids = sequence_padding(batch_segment_ids)
                batch_labels = sequence_padding(batch_labels)
                yield [batch_token_ids, batch_segment_ids], batch_labels
                batch_token_ids, batch_segment_ids, batch_labels = [], [], []


class SIMI:
    maxlen = 128
    batch_size = 20
    config_path = config_path
    checkpoint_path = checkpoint_path
    dict_path = vocab_path
    save_path = bert_base_model_save_path
    data_path = train_data_path
    graph = None
    model =None
    # 建立分词器
    tokenizer = None
    def __init__(self):
        self.tokenizer = Tokenizer(self.dict_path, do_lower_case=True)
        self.graph = tf.get_default_graph()
        # sess = tf.Session(graph=self.graph)
        # 加载预训练模型
        bert = build_transformer_model(
            config_path=self.config_path,
            checkpoint_path=self.checkpoint_path,
            with_pool=True,
            return_keras_model=False,
        )
        # 添加dropout，起到去除过拟合作用
        output = Dropout(rate=0.1)(bert.model.output)
        # 添加输出层，softmax层，两个单元，分别为0,1的概率
        output = Dense(
            units=2, activation='softmax', kernel_initializer=bert.initializer
        )(output)
        # 搭建模型，输入、输出
        self.model = keras.models.Model(bert.model.input, output)
        # model.summary()

    def load_model(self):
        self.model.load_weights(self.save_path)
        return self.model

    # 读取txt数据，返回数据lists[set(text1, text2, int(label))]
    def read_data(self, input_file):
        """Reads a tab separated value file."""
        file = open(input_file, 'r', encoding='utf-8')
        lines = []
        for line in file.read().split('\n'):
            line = line.split("\t")
            if line==['']:
                break
            text1, text2, label = line[0], line[1], line[2]
            text1 = text1.strip('\n')
            text2 = text2.strip('\n')
            label = label.strip('\n')
            lines.append((text1, text2, int(label)))
        return lines

    def biuld_data(self):
        # 加载数据集
        all_data = self.read_data(self.data_path)
        train_data = all_data[:15000]
        random.shuffle(train_data)
        random.shuffle(all_data)
        valid_data = all_data[:2500]
        test_data = all_data[2500:5000]

        # 转换数据集
        batch_size = self.batch_size
        self.train_generator = DataGenerator(train_data, batch_size)
        self.valid_generator = DataGenerator(valid_data, batch_size)
        self.test_generator = DataGenerator(test_data, batch_size)

    def train(self):
        self.biuld_data()
        self.evaluator = Evaluator(self.model, self.valid_generator, self.test_generator)
        # 编译
        self.model.compile(
            loss='sparse_categorical_crossentropy',
            optimizer=Adam(2e-5),
            metrics=['accuracy'],
        )
        self.model.fit_generator(
            self.train_generator.forfit(),
            steps_per_epoch=len(self.train_generator),
            epochs=1,
            callbacks=[self.evaluator]
        )
        self.model.save_weights(self.save_path)

    # 评测函数
    def evaluate(self, data):
        total, right = 0., 0.
        for x_true, y_true in data:
            with self.graph.as_default():
                y_pred = self.model.predict(x_true).argmax(axis=1)
                y_true = y_true[:, 0]
                total += len(y_true)
                right += (y_true == y_pred).sum()
        return right / total

    def test(self):
        file = open(self.data_path, 'r', encoding='utf-8')
        lines = []
        for line in file.read().split('\n'):
            line = line.split("\t")
            if line==['']:
                break
            text1, text2, label = line[0], line[1], line[2]
            text1 = text1.strip('\n')
            text2 = text2.strip('\n')
            label = label.strip('\n')
            lines.append((text1, text2, int(label)))
        test_data = lines[:5000]
        test_generator = DataGenerator(test_data, self.batch_size)
        self.model.load_weights(self.save_path)
        test_acc = self.evaluate(test_generator)
        print(u'test_acc: %.5f\n' %(test_acc))

    def predict_two(self, text1, text2):
        # 两个句子计算相似度
        token_ids, segment_ids = self.tokenizer.encode(text1, text2)
        token_ids, segment_ids = to_array([token_ids], [segment_ids])
        input_data = [token_ids, segment_ids]
        with self.graph.as_default():
            results = self.model.predict(input_data)
        return results

    def predict_many(self, main_text, texts):
        lines = []
        for text in texts:
            lines.append((main_text, text, None))
        # 多个中找最相似的句子
        test_D = DataGenerator(lines, self.batch_size)
        with self.graph.as_default():
            results = self.model.predict_generator(test_D.__iter__(), steps=len(test_D))
        results = results[:, 1]
        result_index = results.argmax(axis=0)
        return result_index, results[result_index]

    def predict_manyto2(self, main_text, texts):
        lines = []
        for text in texts:
            lines.append((main_text, text, None))
        # 多个中找最相似的句子
        test_D = DataGenerator(lines, self.batch_size)
        with self.graph.as_default():
            results = self.model.predict_generator(test_D.__iter__(), steps=len(test_D))
        results = list(results[:, 1])
        result_index_list = list(map(results.index, heapq.nlargest(2, results)))
        return result_index_list

    def find_false(self):
        file = open(train_data_path, 'r', encoding='utf-8')
        file_w = open(r'E:\workspace\BF\inemail\data\wrong_data.txt', 'w', encoding='utf-8')
        for index, line in enumerate(file.read().split('\n')):
            if index > 6650:
                break
            line = line.split("\t")
            text1, text2, label = line[0], line[1], line[2]
            text1 = text1.strip('\n')
            text2 = text2.strip('\n')
            label = label.strip('\n')
            with self.graph.as_default():
                pre_label = self.predict_two(text1, text2)
            if int(label) != int(pre_label[0]):
                print(text1, text2, str(label), str(pre_label[0]))
                file_w.write(text1 + '  ' + text2 + '   ' + str(label) + '  ' + str(pre_label[0]) + '\n')


def simcse_loss(y_true, y_pred):
    """用于SimCSE训练的loss
    """
    # 构造标签
    idxs = K.arange(0, K.shape(y_pred)[0])
    idxs_1 = idxs[None, :]
    idxs_2 = (idxs + 1 - idxs % 2 * 2)[:, None]
    y_true = K.equal(idxs_1, idxs_2)
    y_true = K.cast(y_true, K.floatx())
    # 计算相似度
    y_pred = K.l2_normalize(y_pred, axis=1)
    similarities = K.dot(y_pred, K.transpose(y_pred))
    similarities = similarities - tf.eye(K.shape(y_pred)[0]) * 1e12
    similarities = similarities * 20
    loss = K.categorical_crossentropy(y_true, similarities, from_logits=True)
    return K.mean(loss)


class data_generator(DataGenerator):
    """训练语料生成器
    """
    def __iter__(self, random=False):
        batch_token_ids = []
        for is_end, token_ids in self.sample(random):
            batch_token_ids.append(token_ids)
            batch_token_ids.append(token_ids)
            if len(batch_token_ids) == self.batch_size * 2 or is_end:
                batch_token_ids = sequence_padding(batch_token_ids)
                batch_segment_ids = np.zeros_like(batch_token_ids)
                batch_labels = np.zeros_like(batch_token_ids[:, :1])
                yield [batch_token_ids, batch_segment_ids], batch_labels
                batch_token_ids = []


class SimCSE():
    def __init__(self):
        pooling = 'cls'
        dropout_rate = float(0.3)
        self.encoder = get_encoder(
            config_path,
            checkpoint_path,
            pooling=pooling,
            dropout_rate=dropout_rate
        )
        self.maxlen = 128
        self.data_path = xlsx_path
        self.save_path = simCSE_model_save_path

    def load_model(self):
        dict_path = vocab_path
        self.tokenizer = get_tokenizer(dict_path)
        self.encoder.load_weights(self.save_path)

    def simCSE_predict_many(self, main_text, texts):
        ps = []
        for text in texts:
            a_token_ids = []
            a_token_ids.append(self.tokenizer.encode(main_text, maxlen=self.maxlen)[0])
            b_token_ids = []
            b_token_ids.append(self.tokenizer.encode(text, maxlen=self.maxlen)[0])
            a_token_ids = sequence_padding(a_token_ids)
            b_token_ids = sequence_padding(b_token_ids)
            a_vecs = self.encoder.predict([a_token_ids,
                                      np.zeros_like(a_token_ids)],
                                     verbose=True)
            b_vecs = self.encoder.predict([b_token_ids,
                                      np.zeros_like(b_token_ids)],
                                     verbose=True)
            a_vecs = l2_normalize(a_vecs)
            b_vecs = l2_normalize(b_vecs)
            sims = (a_vecs * b_vecs).sum(axis=1)
            ps.append(sims[0])
        return ps.index(max(ps)), max(ps)

    def train(self):
        wb = openpyxl.load_workbook(self.data_path)
        ws = wb['neo']
        maxrow = ws.max_row  # 最大行
        all_token_ids = []
        common_paths = [train_path,
                        valid_path,
                        test_path]
        datasets = [load_data(path) for path in common_paths]
        for line in range(2, maxrow + 1):
            sentence = ws.cell(line, 4).value
            all_token_ids.append(tokenizer.encode(sentence, maxlen=self.maxlen)[0])
        for data in datasets:
            a_token_ids, b_token_ids, labels = convert_to_ids(data, tokenizer, self.maxlen)
            all_token_ids.extend(a_token_ids)
            all_token_ids.extend(b_token_ids)
            # all_token_ids.append((a_token_ids, b_token_ids))
            # all_labels.append(labels)
            # train_token_ids.extend(a_token_ids)
            # train_token_ids.extend(b_token_ids)
        train_token_ids = sequence_padding(all_token_ids)
        self.encoder.summary()
        self.encoder.compile(loss=simcse_loss, optimizer=Adam(1e-5))
        train_generator = data_generator(train_token_ids, 64)
        self.encoder.fit(
            train_generator.forfit(), steps_per_epoch=len(train_generator), epochs=10
        )
        self.encoder.save(self.save_path)


if __name__ == '__main__':
    # simi = SIMI()
    # simi.train()
    # simi.test()
    # out = simi.predict_two('如何添加普通附件', '如何上传普通附件')
    # print('两句子相似概率：{}%'.format(round(out[0][1]*100, 2)))

    simCSE = SimCSE()
    simCSE.train()