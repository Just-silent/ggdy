# coding:UTF-8
# author    :Just_silent
# init time :2021/5/21 11:39
# file      :evaluator.py
# IDE       :PyCharm

import openpyxl
from QA.response import Response
from tqdm import tqdm

from .path import *

response = Response()
xlsx_path = xlsx_path
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['neo']
maxrow = ws.max_row  # 最大行

nums = 0
ps = 0
pt = []
pf = []
for line in tqdm(range(2, maxrow+1)):
    ori = ws.cell(line, 4).value
    simi = ws.cell(line, 6).value
    if simi is not None:
        simis  = ws.cell(line, 6).value.split('、')
        for simi in simis:
            nums+=1
            question, p = response.evaluator(simi)
            if ori.__eq__(question):
                ps+=1
                pt.append(p)
            else:
                pf.append(p)
print(pt, min(pt))
print(pf, max(pf))
print(ps/nums)