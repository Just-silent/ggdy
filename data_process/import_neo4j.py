# coding:UTF-8
# author    :Just_silent
# init time :2021/5/12 16:02
# file      :import_neo4j.py
# IDE       :PyCharm

import openpyxl
from tqdm import tqdm
from py2neo import Graph, Node, Relationship, NodeMatcher, RelationshipMatcher, Subgraph

from train.path import *

class Neo4jImport():
    '''创建所有的节点和关系
    '''
    def __init__(self):
        self.graf = Graph(
            "http://172.22.179.237:7474/",
            user="neo4j",
            password="123456"
        )
        self.node_match = NodeMatcher(self.graf)
        self.rel_match = RelationshipMatcher(self.graf)
        self.all_nodes = []
        self.all_rels = []
        pass

    def _create_node(self):
        nodes = []
        for name, label in tqdm(self.all_nodes):
            if list(self.node_match.match(label, name=name))==[]:
                nodes.append(Node(label, name=name))
        if nodes!=[]:
            subgraf = Subgraph(nodes)
            self.graf.create(subgraf)
            print('完成创建节点{}个'.format(len(nodes)))
        pass

    def _create_rel(self):
        relation_ships = []
        for (node1, label1), rel, (node2, label2) in tqdm(self.all_rels):
            if  list(self.node_match.match(label1, name=node1))!=[]:
                node1=list(self.node_match.match(label1, name=node1))[0]
            else:
                node1=Node(label1, name=node1)
            if list(self.node_match.match(label2, name=node2)) != []:
                node2 = list(self.node_match.match(label2, name=node2))[0]
            else:
                node2 = Node(label2, name=node2)
            relation_ship = Relationship(node1, rel, node2)
            relation_ships.append(relation_ship)
        if relation_ships!=[]:
            subgraf = Subgraph(relationships=relation_ships)
            self.graf.create(subgraf)
            print('完成创建关系{}个'.format(len(relation_ships)))
        pass

    def macth_node(self, keyword, intent):
        sql = 'match (:keyword{name:\'' + keyword + '\'})-[:`' + intent + '`]->(m) return m'
        result = self.graf.run(sql).data()
        return result

    def create_gddy(self, path, sheet='neo'):
        '''
        xlsx文件格式说明，默认sheet:neo, 目录则必须包括['system', 'domain1', 'keyword', 'question', 'operate']，domain可以只是4个
        :param path:
        :return:
        '''
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        maxrow = ws.max_row  # 最大行
        maxcol = ws.max_column  # 最大列
        names = [ws.cell(1, i).value for i in range(1, maxcol+1)]
        for line in range(2, maxrow+1):
            element = {
                'system': ws.cell(line, names.index('system')+1).value if 'system' in names else None,
                'domain1': ws.cell(line, names.index('domain1')+1).value if 'domain1' in names else None,
                'domain2': ws.cell(line, names.index('domain2')+1).value if 'domain2' in names else None,
                'domain3': ws.cell(line, names.index('domain3')+1).value if 'domain3' in names else None,
                'domain4': ws.cell(line, names.index('domain4')+1).value if 'domain4' in names else None,
                'keyword': ws.cell(line, names.index('keyword')+1).value if 'keyword' in names else None,
                'question': ws.cell(line, names.index('question')+1).value if 'question' in names else None,
                'operate': ws.cell(line, names.index('operate')+1).value if 'operate' in names else None,
            }
            rule_key = ['system', 'domain1', 'domain2', 'domain2', 'domain4', 'keyword', 'question', 'operate']
            all_rel = []
            all_key = []
            for label in element.keys():
                if (element[label], label) not in self.all_nodes and element[label] is not None:
                    self.all_nodes.append((element[label], label))
            for k in rule_key:
                if k in names:
                    all_key.append(k)
            for i in range(len(all_key)-1):
                all_rel.append(all_key[i]+'_'+all_key[i+1])
            for i in range(len(all_key)-1):
                rel = ((element[all_key[i]], all_key[i]), all_rel[i], (element[all_key[i+1]], all_key[i+1]))
                if rel not in self.all_rels:
                    self.all_rels.append(rel)
        self._create_node()
        self._create_rel()
        pass


if __name__ == '__main__':
    neo4j_import = Neo4jImport()
    xlsx_path = xlsx_path
    neo4j_import.create_gddy(xlsx_path)